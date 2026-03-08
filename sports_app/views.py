import uuid
import string
import random
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.db.models import Q, Count, Sum
from django.http import HttpResponse
from django.core.mail import send_mail
from django.conf import settings
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from .models import (
    Team, Achievement, Opportunity, StudentInterestForm,
    EventRegistration, StudentProfile, Announcement,
    GalleryPhoto, Badge, Certificate, LeaderboardEntry, TeamMember
)


# ─── HELPERS ──────────────────────────────────────────────────────────────────

def check_onboarding(request):
    if request.user.is_staff or request.user.is_superuser:
        return None
    try:
        profile = request.user.student_profile
        if not profile.onboarding_complete:
            step = profile.onboarding_step
            if step == 0:
                return redirect('onboarding_step1')
            elif step == 1:
                return redirect('onboarding_step2')
            elif step == 2:
                return redirect('onboarding_step3')
            elif step == 3:
                return redirect('onboarding_step4')
    except StudentProfile.DoesNotExist:
        return redirect('onboarding_step1')
    return None


def send_email_notification(subject, message, recipient_list):
    try:
        send_mail(
            subject=subject,
            message=message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=recipient_list,
            fail_silently=True,
        )
    except Exception:
        pass


def generate_certificate_id():
    chars = string.ascii_uppercase + string.digits
    return 'PIEMR-' + ''.join(random.choices(chars, k=10))


# ─── PUBLIC VIEWS ─────────────────────────────────────────────────────────────

def home(request):
    featured_achievements = Achievement.objects.filter(is_featured=True)[:3]
    upcoming_events = Opportunity.objects.filter(status='open').order_by('start_date')[:3]
    announcements = Announcement.objects.filter(is_active=True)[:5]
    total_teams = Team.objects.filter(is_active=True).count()
    total_achievements = Achievement.objects.count()
    total_students = StudentProfile.objects.filter(onboarding_complete=True).count()
    gallery_photos = GalleryPhoto.objects.filter(is_featured=True)[:6]
    return render(request, 'sports_app/home.html', {
        'featured_achievements': featured_achievements,
        'upcoming_events': upcoming_events,
        'announcements': announcements,
        'total_teams': total_teams,
        'total_achievements': total_achievements,
        'total_students': total_students,
        'gallery_photos': gallery_photos,
    })


def teams_list(request):
    sport_filter = request.GET.get('sport', '')
    teams = Team.objects.filter(is_active=True)
    if sport_filter:
        teams = teams.filter(sport=sport_filter)
    sports = Team.SPORT_CHOICES
    return render(request, 'sports_app/teams_list.html', {
        'teams': teams,
        'sports': sports,
        'sport_filter': sport_filter,
    })


def team_detail(request, team_id):
    team = get_object_or_404(Team, id=team_id)
    achievements = team.achievements.all()[:5]
    opportunities = team.opportunities.filter(status='open')[:3]
    members = team.team_members.filter(is_active=True)
    gallery = team.gallery_photos.all()[:8]
    is_interested = False
    if request.user.is_authenticated:
        try:
            is_interested = request.user.student_profile.sports_interests.filter(id=team_id).exists()
        except StudentProfile.DoesNotExist:
            pass
    return render(request, 'sports_app/team_detail.html', {
        'team': team,
        'achievements': achievements,
        'opportunities': opportunities,
        'members': members,
        'gallery': gallery,
        'is_interested': is_interested,
    })


def achievements_list(request):
    achievements = Achievement.objects.all()
    sport_filter = request.GET.get('sport', '')
    type_filter = request.GET.get('type', '')
    if sport_filter:
        achievements = achievements.filter(team__sport=sport_filter)
    if type_filter:
        achievements = achievements.filter(achievement_type=type_filter)
    return render(request, 'sports_app/achievements_list.html', {
        'achievements': achievements,
        'sports': Team.SPORT_CHOICES,
        'achievement_types': Achievement.ACHIEVEMENT_TYPES,
        'sport_filter': sport_filter,
        'type_filter': type_filter,
    })


def opportunities_list(request):
    opportunities = Opportunity.objects.filter(status='open').order_by('start_date')
    return render(request, 'sports_app/opportunities_list.html', {
        'opportunities': opportunities,
    })


def opportunity_detail(request, opportunity_id):
    opportunity = get_object_or_404(Opportunity, id=opportunity_id)
    is_registered = False
    registration = None
    if request.user.is_authenticated:
        registration = EventRegistration.objects.filter(
            user=request.user, opportunity=opportunity
        ).first()
        is_registered = registration is not None
    return render(request, 'sports_app/opportunity_detail.html', {
        'opportunity': opportunity,
        'is_registered': is_registered,
        'registration': registration,
    })


def announcements_list(request):
    announcements = Announcement.objects.filter(is_active=True)
    return render(request, 'sports_app/announcements.html', {
        'announcements': announcements,
    })


def gallery(request):
    photos = GalleryPhoto.objects.all()
    category_filter = request.GET.get('category', '')
    team_filter = request.GET.get('team', '')
    if category_filter:
        photos = photos.filter(category=category_filter)
    if team_filter:
        photos = photos.filter(team__id=team_filter)
    teams = Team.objects.filter(is_active=True)
    return render(request, 'sports_app/gallery.html', {
        'photos': photos,
        'categories': GalleryPhoto.CATEGORY_CHOICES,
        'teams': teams,
        'category_filter': category_filter,
        'team_filter': team_filter,
    })


def leaderboard(request):
    sport_filter = request.GET.get('sport', 'all')
    entries = LeaderboardEntry.objects.select_related('user', 'user__student_profile')
    if sport_filter != 'all':
        entries = entries.filter(sport=sport_filter)
    entries = entries.order_by('-points')[:50]
    sports = Team.SPORT_CHOICES
    return render(request, 'sports_app/leaderboard.html', {
        'entries': entries,
        'sports': sports,
        'sport_filter': sport_filter,
    })


def search_results(request):
    query = request.GET.get('q', '').strip()
    teams, achievements, opportunities = [], [], []
    if query:
        teams = Team.objects.filter(
            Q(name__icontains=query) | Q(description__icontains=query) |
            Q(sport__icontains=query) | Q(coach_name__icontains=query)
        ).filter(is_active=True)
        achievements = Achievement.objects.filter(
            Q(title__icontains=query) | Q(description__icontains=query) |
            Q(player_names__icontains=query) | Q(location__icontains=query)
        )
        opportunities = Opportunity.objects.filter(
            Q(title__icontains=query) | Q(description__icontains=query) |
            Q(event_type__icontains=query) | Q(location__icontains=query)
        )
    total = len(list(teams)) + len(list(achievements)) + len(list(opportunities))
    return render(request, 'sports_app/search_results.html', {
        'query': query,
        'teams': teams,
        'achievements': achievements,
        'opportunities': opportunities,
        'total': total,
    })


# ─── AUTH VIEWS ───────────────────────────────────────────────────────────────

def register(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        password_confirm = request.POST.get('password_confirm', '')

        if User.objects.filter(username=username).exists():
            messages.error(request, 'Username already taken. Choose another.')
            return render(request, 'sports_app/register.html')
        if User.objects.filter(email=email).exists():
            messages.error(request, 'Email already registered.')
            return render(request, 'sports_app/register.html')
        if password != password_confirm:
            messages.error(request, 'Passwords do not match.')
            return render(request, 'sports_app/register.html')
        if len(password) < 6:
            messages.error(request, 'Password must be at least 6 characters.')
            return render(request, 'sports_app/register.html')

        user = User.objects.create_user(
            username=username, email=email, password=password,
            first_name=first_name, last_name=last_name
        )
        token = str(uuid.uuid4())
        StudentProfile.objects.create(user=user, email_token=token)

        # Welcome email
        send_email_notification(
            subject='Welcome to PIEMR Sports Portal! 🏆',
            message=f'Hi {first_name}!\n\nWelcome to PIEMR Sports Portal. Please complete your onboarding to get started.\n\nBest regards,\nPIEMR Sports Department',
            recipient_list=[email]
        )

        login(request, user)
        messages.success(request, f'Welcome {first_name}! Please complete your profile.')
        return redirect('onboarding_step1')

    return render(request, 'sports_app/register.html')


def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            next_url = request.GET.get('next', '')
            if next_url:
                return redirect(next_url)
            if user.is_staff:
                return redirect('admin_dashboard')
            return redirect('dashboard')
        else:
            messages.error(request, 'Invalid username or password.')
    return render(request, 'sports_app/login.html')


def logout_view(request):
    logout(request)
    messages.success(request, 'Logged out successfully.')
    return redirect('home')


def verify_email(request, token):
    try:
        profile = StudentProfile.objects.get(email_token=token)
        profile.email_verified = True
        profile.email_token = ''
        profile.save()
        messages.success(request, 'Email verified successfully!')
    except StudentProfile.DoesNotExist:
        messages.error(request, 'Invalid verification link.')
    return redirect('dashboard')


# ─── ONBOARDING ───────────────────────────────────────────────────────────────

@login_required(login_url='login')
def onboarding_step1(request):
    try:
        profile = request.user.student_profile
    except StudentProfile.DoesNotExist:
        profile = StudentProfile.objects.create(user=request.user)

    if request.method == 'POST':
        roll = request.POST.get('roll_number', '').strip()
        existing = StudentProfile.objects.filter(roll_number=roll).exclude(user=request.user).first()
        if existing:
            messages.error(request, 'This roll number is already registered.')
            return render(request, 'sports_app/onboarding/step1.html', {'profile': profile})

        profile.roll_number = roll
        profile.branch = request.POST.get('branch', '').strip()
        profile.year = request.POST.get('year') or None
        profile.phone = request.POST.get('phone', '').strip()
        profile.date_of_birth = request.POST.get('date_of_birth') or None
        profile.gender = request.POST.get('gender', '')
        if 'profile_picture' in request.FILES:
            profile.profile_picture = request.FILES['profile_picture']
        if profile.onboarding_step < 1:
            profile.onboarding_step = 1
        profile.save()
        return redirect('onboarding_step2')

    return render(request, 'sports_app/onboarding/step1.html', {'profile': profile})


@login_required(login_url='login')
def onboarding_step2(request):
    try:
        profile = request.user.student_profile
    except StudentProfile.DoesNotExist:
        return redirect('onboarding_step1')

    if request.method == 'POST':
        profile.past_achievements = request.POST.get('past_achievements', '')
        profile.experience_level = request.POST.get('experience_level', '')
        profile.years_playing = request.POST.get('years_playing') or None
        profile.highest_level = request.POST.get('highest_level', '').strip()
        if 'certificates' in request.FILES:
            profile.certificates = request.FILES['certificates']
        if profile.onboarding_step < 2:
            profile.onboarding_step = 2
        profile.save()
        return redirect('onboarding_step3')

    return render(request, 'sports_app/onboarding/step2.html', {'profile': profile})


@login_required(login_url='login')
def onboarding_step3(request):
    try:
        profile = request.user.student_profile
    except StudentProfile.DoesNotExist:
        return redirect('onboarding_step1')

    teams = Team.objects.filter(is_active=True)
    if request.method == 'POST':
        selected_teams = request.POST.getlist('sports_interests')
        profile.sports_interests.set(selected_teams)
        profile.position_played = request.POST.get('position_played', '').strip()
        profile.availability = request.POST.get('availability', '')
        profile.willing_to_travel = 'willing_to_travel' in request.POST
        if profile.onboarding_step < 3:
            profile.onboarding_step = 3
        profile.save()
        return redirect('onboarding_step4')

    return render(request, 'sports_app/onboarding/step3.html', {
        'profile': profile,
        'teams': teams,
    })


@login_required(login_url='login')
def onboarding_step4(request):
    try:
        profile = request.user.student_profile
    except StudentProfile.DoesNotExist:
        return redirect('onboarding_step1')

    if request.method == 'POST':
        profile.height_cm = request.POST.get('height_cm') or None
        profile.weight_kg = request.POST.get('weight_kg') or None
        profile.fitness_level = request.POST.get('fitness_level', '')
        profile.medical_conditions = request.POST.get('medical_conditions', '')
        profile.onboarding_step = 4
        profile.onboarding_complete = True
        profile.save()
        messages.success(request, '🎉 Profile complete! Welcome to PIEMR Sports Portal!')
        return redirect('dashboard')

    return render(request, 'sports_app/onboarding/step4.html', {'profile': profile})


# ─── STUDENT VIEWS ────────────────────────────────────────────────────────────

@login_required(login_url='login')
def dashboard(request):
    check = check_onboarding(request)
    if check:
        return check

    profile = request.user.student_profile
    registrations = EventRegistration.objects.filter(
        user=request.user
    ).select_related('opportunity').order_by('-registration_date')

    upcoming_events = Opportunity.objects.filter(
        status='open'
    ).order_by('start_date')[:4]

    badges = Badge.objects.filter(user=request.user)[:6]
    certificates = Certificate.objects.filter(user=request.user)[:3]

    steps_done = sum([
        bool(profile.roll_number),
        bool(profile.past_achievements),
        bool(profile.sports_interests.count()),
        bool(profile.height_cm),
    ]) + 1
    completion_percent = int((steps_done / 5) * 100)

    announcements = Announcement.objects.filter(is_active=True)[:3]

    return render(request, 'sports_app/dashboard.html', {
        'profile': profile,
        'registrations': registrations,
        'upcoming_events': upcoming_events,
        'badges': badges,
        'certificates': certificates,
        'completion_percent': completion_percent,
        'announcements': announcements,
    })


@login_required(login_url='login')
def student_profile(request, user_id):
    profile_user = get_object_or_404(User, id=user_id)
    if profile_user != request.user and not request.user.is_staff:
        messages.error(request, 'You can only view your own profile.')
        return redirect('dashboard')
    try:
        profile = profile_user.student_profile
    except StudentProfile.DoesNotExist:
        profile = None

    registrations = EventRegistration.objects.filter(
        user=profile_user,
        status__in=['registered', 'confirmed', 'participated']
    ).select_related('opportunity')[:5]

    badges = Badge.objects.filter(user=profile_user)
    certificates = Certificate.objects.filter(user=profile_user)

    return render(request, 'sports_app/student_profile.html', {
        'profile_user': profile_user,
        'profile': profile,
        'registrations': registrations,
        'badges': badges,
        'certificates': certificates,
    })


@login_required(login_url='login')
def submit_interest_form(request):
    check = check_onboarding(request)
    if check:
        return check

    if request.method == 'POST':
        form_data = {
            'user': request.user,
            'roll_number': request.POST.get('roll_number', ''),
            'name': request.POST.get('name', ''),
            'email': request.POST.get('email', ''),
            'phone': request.POST.get('phone', ''),
            'branch': request.POST.get('branch', ''),
            'year': request.POST.get('year', 1),
            'past_achievements': request.POST.get('past_achievements', ''),
            'experience_level': request.POST.get('experience_level', 'beginner'),
            'additional_comments': request.POST.get('additional_comments', ''),
        }
        interest_form = StudentInterestForm.objects.create(**form_data)
        selected_teams = request.POST.getlist('sports_interests')
        interest_form.sports_interests.set(selected_teams)

        send_email_notification(
            subject='Interest Form Submitted — PIEMR Sports',
            message=f'Hi {request.user.first_name},\n\nYour interest form has been submitted successfully. Our coaches will review it shortly.\n\nPIEMR Sports Department',
            recipient_list=[request.user.email]
        )

        messages.success(request, '✅ Interest form submitted! Coaches will review it soon.')
        return redirect('dashboard')

    profile = request.user.student_profile
    teams = Team.objects.filter(is_active=True)
    return render(request, 'sports_app/submit_interest_form.html', {
        'profile': profile,
        'teams': teams,
    })


@login_required(login_url='login')
def register_event(request, opportunity_id):
    check = check_onboarding(request)
    if check:
        return check

    opportunity = get_object_or_404(Opportunity, id=opportunity_id)

    if EventRegistration.objects.filter(user=request.user, opportunity=opportunity).exists():
        messages.warning(request, 'You are already registered for this event.')
        return redirect('opportunity_detail', opportunity_id=opportunity_id)

    if request.method == 'POST':
        EventRegistration.objects.create(
            user=request.user,
            opportunity=opportunity,
            team_name=request.POST.get('team_name', ''),
            team_members=request.POST.get('team_members', ''),
            contact_number=request.POST.get('contact_number', ''),
        )

        # Update leaderboard points
        entry, created = LeaderboardEntry.objects.get_or_create(
            user=request.user,
            sport=opportunity.team.sport,
            defaults={'points': 0}
        )
        entry.points += 10
        entry.events_participated += 1
        entry.save()

        # Award participation badge
        Badge.objects.create(
            user=request.user,
            badge_type='participation',
            title=f'Registered for {opportunity.title}',
            sport=opportunity.team.sport,
            event=opportunity
        )

        # Issue certificate
        Certificate.objects.create(
            user=request.user,
            title=f'Participation Certificate — {opportunity.title}',
            event=opportunity,
            certificate_id=generate_certificate_id()
        )

        send_email_notification(
            subject=f'Registration Confirmed — {opportunity.title}',
            message=f'Hi {request.user.first_name},\n\nYou have successfully registered for {opportunity.title}.\n\nEvent Date: {opportunity.start_date.strftime("%d %B %Y")}\nVenue: {opportunity.location}\n\nBest of luck!\nPIEMR Sports Department',
            recipient_list=[request.user.email]
        )

        messages.success(request, f'✅ Successfully registered for {opportunity.title}!')
        return redirect('dashboard')

    return render(request, 'sports_app/register_event.html', {
        'opportunity': opportunity,
    })


@login_required(login_url='login')
def cancel_registration(request, registration_id):
    registration = get_object_or_404(EventRegistration, id=registration_id, user=request.user)
    registration.status = 'cancelled'
    registration.save()
    messages.success(request, 'Registration cancelled.')
    return redirect('dashboard')


# ─── ADMIN VIEWS ──────────────────────────────────────────────────────────────

@login_required(login_url='login')
def admin_dashboard(request):
    if not request.user.is_staff:
        messages.error(request, 'Access denied.')
        return redirect('home')

    total_students = StudentProfile.objects.filter(onboarding_complete=True).count()
    total_teams = Team.objects.filter(is_active=True).count()
    total_registrations = EventRegistration.objects.count()
    total_achievements = Achievement.objects.count()
    open_events = Opportunity.objects.filter(status='open').count()
    pending_forms = StudentInterestForm.objects.filter(is_reviewed=False).count()
    total_badges = Badge.objects.count()
    total_certificates = Certificate.objects.count()

    interest_forms = StudentInterestForm.objects.select_related('user').order_by('-submission_date')[:20]
    registrations = EventRegistration.objects.select_related('user', 'opportunity').order_by('-registration_date')[:20]
    recent_students = StudentProfile.objects.filter(
        onboarding_complete=True
    ).select_related('user').order_by('-created_at')[:10]

    announcements = Announcement.objects.all()[:5]

    # Chart data — registrations per sport
    sport_data = EventRegistration.objects.values(
        'opportunity__team__sport'
    ).annotate(count=Count('id')).order_by('-count')

    # Recent activity
    recent_registrations = EventRegistration.objects.select_related(
        'user', 'opportunity'
    ).order_by('-registration_date')[:5]

    return render(request, 'sports_app/admin_dashboard.html', {
        'total_students': total_students,
        'total_teams': total_teams,
        'total_registrations': total_registrations,
        'total_achievements': total_achievements,
        'open_events': open_events,
        'pending_forms': pending_forms,
        'total_badges': total_badges,
        'total_certificates': total_certificates,
        'interest_forms': interest_forms,
        'registrations': registrations,
        'recent_students': recent_students,
        'announcements': announcements,
        'sport_data': list(sport_data),
        'recent_registrations': recent_registrations,
    })


@login_required(login_url='login')
def award_badge(request):
    if not request.user.is_staff:
        return redirect('home')
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        badge_type = request.POST.get('badge_type')
        title = request.POST.get('title')
        sport = request.POST.get('sport', '')
        user = get_object_or_404(User, id=user_id)
        Badge.objects.create(
            user=user, badge_type=badge_type,
            title=title, sport=sport,
            awarded_by=request.user
        )
        # Update leaderboard
        if sport:
            entry, _ = LeaderboardEntry.objects.get_or_create(
                user=user, sport=sport, defaults={'points': 0}
            )
            points_map = {'gold': 100, 'silver': 75, 'bronze': 50, 'excellence': 30, 'leadership': 20, 'participation': 10}
            entry.points += points_map.get(badge_type, 10)
            entry.medals += 1 if badge_type in ['gold', 'silver', 'bronze'] else 0
            entry.save()
        messages.success(request, f'Badge awarded to {user.get_full_name()}!')
    return redirect('admin_dashboard')


@login_required(login_url='login')
def issue_certificate(request):
    if not request.user.is_staff:
        return redirect('home')
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        title = request.POST.get('title')
        user = get_object_or_404(User, id=user_id)
        Certificate.objects.create(
            user=user, title=title,
            certificate_id=generate_certificate_id()
        )
        messages.success(request, f'Certificate issued to {user.get_full_name()}!')
    return redirect('admin_dashboard')


# ─── EXCEL EXPORTS ────────────────────────────────────────────────────────────

def style_header(ws, headers, fill_color='1a3c6e'):
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    font = Font(bold=True, color='FFFFFF', size=11)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center')


@login_required(login_url='login')
def export_interest_forms_excel(request):
    if not request.user.is_staff:
        return redirect('home')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Interest Forms'
    headers = ['Name', 'Roll No', 'Branch', 'Year', 'Email', 'Phone', 'Sports', 'Experience', 'Submitted', 'Reviewed']
    style_header(ws, headers)
    for form in StudentInterestForm.objects.all():
        ws.append([
            form.name, form.roll_number, form.branch, form.year,
            form.email, form.phone,
            ', '.join([t.name for t in form.sports_interests.all()]),
            form.experience_level, str(form.submission_date.date()),
            'Yes' if form.is_reviewed else 'No'
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="interest_forms.xlsx"'
    wb.save(response)
    return response


@login_required(login_url='login')
def export_registrations_excel(request):
    if not request.user.is_staff:
        return redirect('home')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Registrations'
    headers = ['Student', 'Roll No', 'Event', 'Sport', 'Date', 'Status', 'Contact']
    style_header(ws, headers)
    for reg in EventRegistration.objects.select_related('user', 'opportunity', 'user__student_profile').all():
        roll = ''
        try:
            roll = reg.user.student_profile.roll_number or ''
        except Exception:
            pass
        ws.append([
            reg.user.get_full_name(), roll,
            reg.opportunity.title, reg.opportunity.team.sport,
            str(reg.registration_date.date()), reg.status,
            reg.contact_number
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="registrations.xlsx"'
    wb.save(response)
    return response


@login_required(login_url='login')
def export_student_profiles_excel(request):
    if not request.user.is_staff:
        return redirect('home')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Student Profiles'
    headers = ['Name', 'Username', 'Email', 'Roll No', 'Branch', 'Year', 'Phone', 'Experience', 'Sports', 'Joined']
    style_header(ws, headers)
    for profile in StudentProfile.objects.select_related('user').filter(onboarding_complete=True):
        ws.append([
            profile.user.get_full_name(), profile.user.username,
            profile.user.email, profile.roll_number or '',
            profile.branch or '', profile.year or '',
            profile.phone or '', profile.experience_level or '',
            ', '.join([t.name for t in profile.sports_interests.all()]),
            str(profile.created_at.date())
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="student_profiles.xlsx"'
    wb.save(response)
    return response